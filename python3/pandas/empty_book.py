#!/usr/bin/python3
# coding: utf-8

''' xlsx practice '''

import sys
try:
    from openpyxl import Workbook
    #from openpyxl.utils import get_column_letter
except ImportError:
    print('cannot import openpyxl, quit...')
    sys.exit(1)

def main():
    ''' main '''
    wb = Workbook()
    fn = 'empty_book.xlsx'
    ws1 = wb.active
    ws1.title = "multiple"
    for row in range(1, 20):
        for col in range(1, 20):
            v = row * col
            ws1.cell(column=col, row=row, value=v)

    wb.save(filename=fn)

if __name__ == '__main__':
    main()
