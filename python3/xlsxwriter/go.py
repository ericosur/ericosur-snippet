#!/usr/bin/python3
# coding: utf-8

''' read from and write into '''

# pylint: disable=import-error
from time import sleep
from openpyxl import load_workbook

class Solution():
    ''' solution to combine xlsx '''
    def __init__(self):
        self.sheet_name = 'translation_CIS008_Plus_2020012'
        self.src_xlsx = 'src.xlsx'
        self.trans_dict = {}
        self.src_dict = {}

    @staticmethod
    def getcn(col, row):
        ''' get cell name '''
        cn = f'{col}{row}'
        return cn


    def read_xlsx(self, fn, cfrom, cto, col_initial):
        ''' read_xlsx '''
        wb_ar = load_workbook(filename=fn)
        ar = wb_ar[self.sheet_name]
        # wb_src = load_workbook(filename='src.xlsx')
        # src = wb_src[SHEET_NAME]

        mydict = {}
        for i in range(cfrom, cto):
            myvalues = tuple()
            acol = Solution.getcn('A', i)
            dcol = Solution.getcn(col_initial, i)
            k = ar[acol].value
            if isinstance(k, str):
                k = int(k)
            v = ar[dcol].value
            v = v.strip()
            myvalues = (dcol, v)
            mydict[k] = myvalues
        return mydict


    @staticmethod
    def output(data_dict):
        ''' output '''
        for kk in sorted(data_dict.keys()):
            vv = data_dict[kk]
            print(f'[{kk}]__{vv}__')

    def check(self, dict1, src_dict):
        ''' compare values '''
        wb = load_workbook(filename=self.src_xlsx)
        sh = wb[self.sheet_name]
        no_same = 0
        no_diff = 0
        for kk, vv in dict1.items():
            if kk in src_dict:
                src_vv = src_dict[kk]
            else:
                print("[ERROR] key not exist: ", kk)
            if vv[1] != src_vv[1]:
                no_diff += 1
                #print('___新____{}\n舊_{}__{}\n'.format(vv[1], src_vv[0], src_vv[1]))
                sh[src_vv[0]] = vv[1]
            else:
                no_same += 1
        print(f'diff({no_diff}), same({no_same}), save xlsx...')
        wb.save(filename=self.src_xlsx)
        #wb.save(filename='new.xlsx')

    def run(self):
        ''' run '''
        fn = 'Arabic.XLSX'
        print(f'process {fn}...')
        self.trans_dict = self.read_xlsx(fn, 2, 47, 'D')
        #self.output(self.ar_dict)
        self.src_dict = self.read_xlsx(self.src_xlsx, 2, 448, 'D')
        #self.output(self.src_dict)
        self.check(self.trans_dict, self.src_dict)
        sleep(1)

        fn = 'RUssian.XLSX'
        print(f'process {fn}...')
        self.trans_dict = self.read_xlsx(fn, 2, 47, 'G')
        self.src_dict = self.read_xlsx(self.src_xlsx, 2, 448, 'G')
        self.check(self.trans_dict, self.src_dict)
        sleep(1)

        fn = 'Portuguese and Spanish.XLSX'
        print(f'process {fn}...')
        self.trans_dict = self.read_xlsx(fn, 2, 47, 'E')
        self.src_dict = self.read_xlsx(self.src_xlsx, 2, 448, 'E')
        self.check(self.trans_dict, self.src_dict)
        sleep(1)
        self.trans_dict = self.read_xlsx(fn, 2, 47, 'F')
        self.src_dict = self.read_xlsx(self.src_xlsx, 2, 448, 'F')
        self.check(self.trans_dict, self.src_dict)


def main():
    ''' main '''
    go = Solution()
    go.run()


if __name__ == '__main__':
    main()
